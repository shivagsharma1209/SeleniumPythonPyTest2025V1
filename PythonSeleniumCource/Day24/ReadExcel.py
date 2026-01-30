import openpyxl

#File-->Workbook-->Sheets-->Rows-->cells
file = 'C:/Users/LENOVO/OneDrive/Documents/PythonSelenium/PythonDDT/DataDrivenTest1.xlsx'
wb = openpyxl.load_workbook(file)
sheet = wb['Sheet1']

rows = sheet.max_row    # number of roes in a sheet
cols = sheet.max_column # number of columns in a sheet

for row in range(1, rows+1):
    for col in range(1, cols+1):
        print(sheet.cell(row, col).value,end ='        ')
    print()







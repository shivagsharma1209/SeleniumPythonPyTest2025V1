import openpyxl
from openpyxl.styles import PatternFill

workbook = openpyxl.Workbook()
sheet = workbook.active

def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.cell(row=rownum, column=columnnum).value)

def writeData(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.cell(row=rownum, column=columnnum).value)
    workbook.save(file)

def fillGreenColor(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    fillGreen = PatternFill(start_color='60b212',end_color ='60b212',fill_type='solid')
    sheet.cell(rownum,columnnum).fill = fillGreen
    workbook.save(file)

def fillRedColor(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    fillRed = PatternFill(start_color='ff0000',end_color = 'ff0000',fill_type='solid')
    sheet.cell(rownum,columnnum).fill = fillRed
    workbook.save(file)







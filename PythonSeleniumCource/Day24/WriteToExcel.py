import openpyxl

# 1. To write the same data in all cells
file = 'C:/Users/LENOVO/OneDrive/Documents/PythonSelenium/PythonDDT/DataDrivenTest2.xlsx'

wb = openpyxl.load_workbook(file)
sheet = wb.active #To get the active sheet from the excel sheet

for row in range(1, 17):
    for col in range(1, 10):
        sheet.cell(row,col).value = "Welcome To Python"
wb.save(file)

#2. To write multiple sets of data.........

file = 'C:/Users/LENOVO/OneDrive/Documents/PythonSelenium/PythonDDT/DataDrivenTest3.xlsx'

wb = openpyxl.load_workbook(file)
sheet = wb.active #To get the active sheet from the excel sheet

sheet.cell(1,1).value = 'Name'
sheet.cell(1,2).value = 'Emp Id'
sheet.cell(1,3).value = 'Designation'
sheet.cell(1,4).value = 'Salary'

sheet.cell(2,1).value = 'Shivaprasad'
sheet.cell(2,2).value = 10001
sheet.cell(2,3).value = 'Associate'
sheet.cell(2,4).value = 15000


sheet.cell(3,1).value = 'Krishna'
sheet.cell(3,2).value = 10002
sheet.cell(3,3).value = 'Director'
sheet.cell(3,4).value = 50000

sheet.cell(4,1).value = 'Rama Krishna'
sheet.cell(4,2).value = 10002
sheet.cell(4,3).value = 'Manager'
sheet.cell(4,4).value = 30000

wb.save(file)





